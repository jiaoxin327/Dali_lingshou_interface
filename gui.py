from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                           QPushButton, QTextEdit, QLabel, QMessageBox, QLineEdit,
                           QFormLayout, QTabWidget, QGroupBox, QTimeEdit, QCheckBox,
                           QTableWidget, QTableWidgetItem, QHeaderView, QDialog, QComboBox,
                           QInputDialog, QSizeGrip, QFileDialog, QProgressBar, QDialogButtonBox)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QTime
from retail_api import RetailAPI
from db_utils import DatabaseConnection
from utils.logger import Logger
from utils.validator import DataValidator
import sys
import json
import os
import schedule
import time
from datetime import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font

class WorkerThread(QThread):
    """后台工作线程，避免界面卡顿"""
    update_signal = pyqtSignal(str)  # 用于更新界面的信号
    finished_signal = pyqtSignal(bool, str)  # 完成信号，带状态和消息
    refresh_history_signal = pyqtSignal()  # 添加刷新历史信号

    def save_history(self, status, data_count, message, error_detail=None):
        """保存上报历史到JSON文件"""
        try:
            history_file = 'upload_history.json'
            history_data = []
            
            # 读取现有历史记录
            if os.path.exists(history_file):
                try:
                    with open(history_file, 'r', encoding='utf-8') as f:
                        history_data = json.load(f)
                except:
                    history_data = []
            
            # 添加新记录
            new_record = {
                'upload_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'status': status,
                'data_count': data_count,
                'message': message,
                'error_detail': error_detail,
                'source': '接口导入'  # 添加数据来源标识
            }
            
            # 将新记录添加到开头
            history_data.insert(0, new_record)
            
            # 只保留最近100条记录
            history_data = history_data[:100]
            
            # 保存到文件
            with open(history_file, 'w', encoding='utf-8') as f:
                json.dump(history_data, f, indent=4, ensure_ascii=False)
                
        except Exception as e:
            print(f"保存历史记录失败: {str(e)}")
    
    def run(self):
        try:
            self.update_signal.emit("开始执行数据上报...")
            
            # 加载配置
            with open('config.json', 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            # 初始化API客户端
            api = RetailAPI(config['api']['url'])
            
            # 登录系统
            self.update_signal.emit("正在登录系统...")
            if not api.login(config['api']['username'], config['api']['password']):
                self.finished_signal.emit(False, "登录失败")
                return
            
            self.update_signal.emit("正在获取数据...")
            # 获取数据
            db = DatabaseConnection(**config['database'])
            
            if not db.test_connection():
                self.finished_signal.emit(False, "数据库连接失败")
                return
                
            if not db.check_table_exists():
                self.finished_signal.emit(False, "数据表不存在")
                return
                
            data = db.get_retail_data()
            
            # 数据验证
            self.update_signal.emit("正在验证数据...")
            failed_records = DataValidator.validate_batch_data(data)
            if failed_records:
                error_msg = "数据验证失败:\n"
                for record in failed_records:
                    error_msg += f"数据: {record['data']}\n错误: {record['error']}\n"
                self.finished_signal.emit(False, error_msg)
                return
                
            if not data:
                self.finished_signal.emit(False, "没有获取到需要上报的数据")
                return
                
            # 上报数据
            self.update_signal.emit("正在上报数据...")
            result = api.upload_retail_data(data)
            if result and result.get("code") == 200:
                success_msg = "数据上报成功\n"
                for item in result.get("content", []):
                    success_msg += f"数据ID: {item['soureId']}, 状态: {item['code']}, 消息: {item['msg']}\n"
                self.finished_signal.emit(True, success_msg)
                # 保存成功历史
                self.save_history(
                    status='成功',
                    data_count=len(data),
                    message=str(result.get("content", [])),
                    error_detail=None
                )
                # 发送刷新历史信号
                self.refresh_history_signal.emit()
            else:
                self.finished_signal.emit(False, f"数据上报失败: {str(result)}")
                # 保存失败历史
                self.save_history(
                    status='失败',
                    data_count=len(data),
                    message="上报失败",
                    error_detail=str(result)
                )
                
        except Exception as e:
            self.finished_signal.emit(False, f"执行出错: {str(e)}")
            # 保存错误历史
            self.save_history(
                status='失败',
                data_count=0,
                message="执行出错",
                error_detail=str(e)
            )
        finally:
            if 'db' in locals():
                db.close()

class ConfigTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.initUI()
        self.loadConfig()
        
    def initUI(self):
        layout = QVBoxLayout()
        
        # 数据库配置组
        db_group = QGroupBox("数据库配置")
        db_layout = QFormLayout()
        
        self.db_host = QLineEdit()
        self.db_port = QLineEdit()
        self.db_user = QLineEdit()
        self.db_password = QLineEdit()
        self.db_name = QLineEdit()
        
        self.db_password.setEchoMode(QLineEdit.Password)
        
        db_layout.addRow("主机:", self.db_host)
        db_layout.addRow("端口:", self.db_port)
        db_layout.addRow("用户名:", self.db_user)
        db_layout.addRow("密码:", self.db_password)
        db_layout.addRow("数据库名:", self.db_name)
        
        # 添加数据测试按钮
        db_test_button = QPushButton("测试数据库连接")
        db_test_button.clicked.connect(self.test_db_connection)
        db_test_button.setStyleSheet('''
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        ''')
        db_layout.addRow("", db_test_button)
        
        db_group.setLayout(db_layout)
        
        # API配置组
        api_group = QGroupBox("API配置")
        api_layout = QFormLayout()
        
        self.api_url = QLineEdit()
        self.api_username = QLineEdit()
        self.api_password = QLineEdit()
        
        self.api_password.setEchoMode(QLineEdit.Password)
        
        api_layout.addRow("API地址:", self.api_url)
        api_layout.addRow("用户名:", self.api_username)
        api_layout.addRow("密码:", self.api_password)
        
        # 添加API测试按钮
        api_test_button = QPushButton("测试API连接")
        api_test_button.clicked.connect(self.test_api_connection)
        api_test_button.setStyleSheet('''
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        ''')
        api_layout.addRow("", api_test_button)
        
        api_group.setLayout(api_layout)
        
        # 按钮布局
        button_layout = QHBoxLayout()
        
        # 保存按��
        save_button = QPushButton("保存配置")
        save_button.clicked.connect(self.saveConfig)
        save_button.setStyleSheet('''
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        ''')
        
        # 测试全部按钮
        test_all_button = QPushButton("测试全部连接")
        test_all_button.clicked.connect(self.test_all_connections)
        test_all_button.setStyleSheet('''
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        ''')
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(test_all_button)
        
        layout.addWidget(db_group)
        layout.addWidget(api_group)
        layout.addLayout(button_layout)
        layout.addStretch()
        
        self.setLayout(layout)
        
    def test_db_connection(self):
        """测试数据库连接"""
        try:
            # 确保端口号是整数
            try:
                port = int(self.db_port.text()) if self.db_port.text() else 3306
            except ValueError:
                QMessageBox.warning(self, "错误", "端口号必须是数字！")
                return
            
            db_config = {
                'host': self.db_host.text(),
                'port': port,
                'user': self.db_user.text(),
                'password': self.db_password.text(),
                'database': self.db_name.text()
            }
            
            db = DatabaseConnection(**db_config)
            if db.test_connection():
                QMessageBox.information(self, "成功", "数据库连接测试成功！")
            else:
                QMessageBox.warning(self, "错误", "数据库连接测试失败！")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"数据库连接测试出错：{str(e)}")
            
    def test_api_connection(self):
        """测试API连接"""
        try:
            api = RetailAPI(self.api_url.text())
            if api.login(self.api_username.text(), self.api_password.text()):
                QMessageBox.information(self, "成功", "API连接测试成功！")
            else:
                QMessageBox.warning(self, "错误", "API连接测试失败！")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"API连接测试出错：{str(e)}")
            
    def test_all_connections(self):
        """测试所有连接"""
        # 测试数据库连接
        try:
            db_config = {
                'host': self.db_host.text(),
                'port': int(self.db_port.text() or 3306),
                'user': self.db_user.text(),
                'password': self.db_password.text(),
                'database': self.db_name.text()
            }
            
            db = DatabaseConnection(**db_config)
            db_success = db.test_connection()
        except Exception as e:
            db_success = False
            db_error = str(e)
            
        # 测试API连接
        try:
            api = RetailAPI(self.api_url.text())
            api_success = api.login(self.api_username.text(), self.api_password.text())
        except Exception as e:
            api_success = False
            api_error = str(e)
            
        # 显示测试结果
        result_message = "连接测试结果：\n\n"
        result_message += f"数据库连接: {'成功' if db_success else '失败'}\n"
        if not db_success:
            result_message += f"数据库错误: {db_error if 'db_error' in locals() else '连接失败'}\n\n"
            
        result_message += f"API连接: {'成功' if api_success else '失败'}\n"
        if not api_success:
            result_message += f"API错误: {api_error if 'api_error' in locals() else '连接失败'}\n"
            
        if db_success and api_success:
            QMessageBox.information(self, "成功", "所有连接测试成功！")
        else:
            QMessageBox.warning(self, "错误", result_message)
        
    def loadConfig(self):
        """加载配置文件"""
        config_file = 'config.json'
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    
                # 设置数据库配置
                db_config = config.get('database', {})
                self.db_host.setText(db_config.get('host', 'localhost'))
                self.db_port.setText(str(db_config.get('port', '3306')))
                self.db_user.setText(db_config.get('user', 'root'))
                self.db_password.setText(db_config.get('password', ''))
                self.db_name.setText(db_config.get('database', 'retail_report'))
                
                # 设置API配置
                api_config = config.get('api', {})
                self.api_url.setText(api_config.get('url', 'http://49.235.172.155:3727/supply-security-api'))
                self.api_username.setText(api_config.get('username', 'SFJRPA1234'))
                self.api_password.setText(api_config.get('password', ''))
            except Exception as e:
                QMessageBox.warning(self, "错误", f"加载配置文件失败: {str(e)}")
                
    def saveConfig(self):
        """保存配置到文件"""
        config = {
            'database': {
                'host': self.db_host.text(),
                'port': int(self.db_port.text() or 3306),
                'user': self.db_user.text(),
                'password': self.db_password.text(),
                'database': self.db_name.text()
            },
            'api': {
                'url': self.api_url.text(),
                'username': self.api_username.text(),
                'password': self.api_password.text()
            }
        }
        
        try:
            with open('config.json', 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
            QMessageBox.information(self, "成功", "配置保存成功！")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存配置文件失败: {str(e)}")

class ScheduleTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.timer = None
        self.schedule_job = None
        self.main_window = None
        self.initUI()
        
    def set_main_window(self, main_window):
        """���置主窗口引用"""
        self.main_window = main_window
        
    def initUI(self):
        layout = QVBoxLayout()
        
        # 定时任务配置组
        schedule_group = QGroupBox("���配置")
        schedule_layout = QFormLayout()
        
        # 启用定时任务复选框
        self.enable_schedule = QCheckBox("启用定时任务")
        self.enable_schedule.stateChanged.connect(self.on_schedule_changed)
        schedule_layout.addRow(self.enable_schedule)
        
        # 时间选择器
        self.time_edit = QTimeEdit()
        self.time_edit.setDisplayFormat("HH:mm")
        self.time_edit.setTime(QTime(9, 0))  # 默认设置为早上9点
        schedule_layout.addRow("执行时间:", self.time_edit)
        
        # 状态显示
        self.status_label = QLabel("定时任务未启动")
        self.status_label.setStyleSheet("color: gray;")
        schedule_layout.addRow("当前状态:", self.status_label)
        
        # 下次执行时间显示
        self.next_run_label = QLabel("-")
        schedule_layout.addRow("下次执行:", self.next_run_label)
        
        schedule_group.setLayout(schedule_layout)
        
        # 控制按钮
        button_layout = QHBoxLayout()
        
        self.start_button = QPushButton("启动任务")
        self.start_button.clicked.connect(self.start_schedule)
        self.start_button.setStyleSheet('''
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        ''')
        
        self.stop_button = QPushButton("停止任务")
        self.stop_button.clicked.connect(self.stop_schedule)
        self.stop_button.setEnabled(False)
        self.stop_button.setStyleSheet('''
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        ''')
        
        button_layout.addWidget(self.start_button)
        button_layout.addWidget(self.stop_button)
        
        layout.addWidget(schedule_group)
        layout.addLayout(button_layout)
        layout.addStretch()
        
        self.setLayout(layout)
        
    def on_schedule_changed(self, state):
        """定时任务启用状态改变"""
        self.time_edit.setEnabled(state == Qt.Checked)
        self.start_button.setEnabled(state == Qt.Checked)
        
    def start_schedule(self):
        """启动定时任务"""
        if not self.enable_schedule.isChecked():
            return
            
        if not self.main_window:
            QMessageBox.warning(self, "错误", "未找到主窗口引用")
            return
            
        schedule_time = self.time_edit.time()
        hour = schedule_time.hour()
        minute = schedule_time.minute()
        
        # 清除之前的任务
        schedule.clear()
        
        # 设置新任务
        schedule.every().day.at(f"{hour:02d}:{minute:02d}").do(self.main_window.start_upload)
        
        # 启动定时器检查任务
        if not self.timer:
            self.timer = QTimer()
            self.timer.timeout.connect(self.check_schedule)
        self.timer.start(1000)  # 每秒检查一次
        
        self.status_label.setText("定时任务已启动")
        self.status_label.setStyleSheet("color: green;")
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.time_edit.setEnabled(False)
        self.enable_schedule.setEnabled(False)
        
        self.update_next_run()
        
    def stop_schedule(self):
        """停止定时任务"""
        if self.timer:
            self.timer.stop()
        schedule.clear()
        
        self.status_label.setText("定时任务已停止")
        self.status_label.setStyleSheet("color: red;")
        self.next_run_label.setText("-")
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.time_edit.setEnabled(True)
        self.enable_schedule.setEnabled(True)
        
    def check_schedule(self):
        """检查并运行定时任务"""
        try:
            schedule.run_pending()
            self.update_next_run()
        except Exception as e:
            self.status_label.setText(f"定时任务执行出错: {str(e)}")
            self.status_label.setStyleSheet("color: red;")
            QMessageBox.warning(self, "错误", f"定时任务执行出错: {str(e)}")
        
    def update_next_run(self):
        """更新下次运行时间显示"""
        next_run = schedule.next_run()
        if next_run:
            self.next_run_label.setText(next_run.strftime("%Y-%m-%d %H:%M:%S"))
        else:
            self.next_run_label.setText("-")

class HistoryTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.initUI()
        
    def initUI(self):
        layout = QVBoxLayout()
        
        # 添加标题
        title_label = QLabel('上报历史记录')
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet('font-size: 16px; font-weight: bold; margin: 10px;')
        layout.addWidget(title_label)
        
        # 创建表格
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(['上报时间', '状态', '数据条数', '结果消息', '错误详情', '数据来源', '记录时间'])
        
        # 设置表格样式
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet('''
            QTableWidget {
                gridline-color: #d0d0d0;
                background-color: white;
                alternate-background-color: #f6f6f6;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 4px;
                border: 1px solid #d0d0d0;
                font-weight: bold;
            }
        ''')
        
        layout.addWidget(self.table)
        
        # 添加刷新按钮
        refresh_button = QPushButton('刷新历史记录')
        refresh_button.clicked.connect(self.refresh_history)
        refresh_button.setStyleSheet('''
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        ''')
        layout.addWidget(refresh_button, alignment=Qt.AlignRight)
        
        self.setLayout(layout)
        
        # 初始加载数据
        self.refresh_history()
        
    def refresh_history(self):
        """刷新历史记录"""
        try:
            history_file = 'upload_history.json'
            if not os.path.exists(history_file):
                return
                
            with open(history_file, 'r', encoding='utf-8') as f:
                history_data = json.load(f)
            
            # 清空表格
            self.table.setRowCount(0)
            
            # 填充数据
            for row_num, row in enumerate(history_data):
                self.table.insertRow(row_num)
                
                # 设置单元格内容
                self.table.setItem(row_num, 0, QTableWidgetItem(row['upload_time']))
                
                status_item = QTableWidgetItem(row['status'])
                status_item.setForeground(Qt.green if row['status'] == '成功' else Qt.red)
                self.table.setItem(row_num, 1, status_item)
                
                self.table.setItem(row_num, 2, QTableWidgetItem(str(row['data_count'])))
                self.table.setItem(row_num, 3, QTableWidgetItem(row['message']))
                self.table.setItem(row_num, 4, QTableWidgetItem(row['error_detail'] or ''))
                self.table.setItem(row_num, 5, QTableWidgetItem(row.get('source', '接口导入')))  # 显示数据来源
                
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载历史记录失败: {str(e)}")

class TableMappingTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        # 添加历史映射配置存储
        self.mapping_history_file = 'mapping_history.json'
        self.mapping_history = self.load_mapping_history()
        self.default_mappings = [
            ('social_credit_code', 'socialCreditCode', '统一社会信用代码'),
            ('comp_name', 'compName', '企业名称'),
            ('retail_store_code', 'retailStoreCode', '零售点编码'),
            ('retail_store_name', 'retailStoreName', '零售点名称'),
            ('report_date', 'reportDate', '上报日期'),
            ('commodity_code', 'selfCommondityCode', '商品编码'),
            ('commodity_name', 'selfCommondityName', '商品名称'),
            ('unit', 'unit', '单位'),
            ('spec', 'spec', '规格'),
            ('barcode', 'barcode', '条形码'),
            ('data_type', 'dataType', '数据类型'),
            ('data_value', 'dataValue', '数据值'),
            ('data_convert_flag', 'dataConvertFlag', '数据转换标志'),
            ('standard_commodity_code', 'standardCommondityCode', '标准商品编码'),
            ('standard_commodity_name', 'standardCommondityName', '标准商品名称'),
            ('package_name', 'packageName', '包装名称'),
            ('supplier_code', 'supplierCode', '供应商编码'),
            ('supplier_name', 'supplierName', '供应商名称'),
            ('manufacturer', 'manufatureName', '生产厂家'),
            ('origin_code', 'originCode', '产地编码'),
            ('origin_name', 'originName', '产地名称'),
            ('scene_flag', 'sceneflag', '场景标志')
        ]
        self.initUI()
        self.load_default_config()  # 替换原来的自动重置为加载默认配置
        
    def load_default_config(self):
        """静默加载默认配置"""
        # 设置默认表名
        self.table_name.setText('retail_data')
        
        # 清空当前表格
        self.mapping_table.setRowCount(0)
        
        # 添加默认映射
        for i, (db_field, api_field, description) in enumerate(self.default_mappings):
            self.mapping_table.insertRow(i)
            self.mapping_table.setItem(i, 0, QTableWidgetItem(db_field))
            self.mapping_table.setItem(i, 1, QTableWidgetItem(api_field))
            self.mapping_table.setItem(i, 2, QTableWidgetItem(description))
            
        # 静默保存默认配置
        try:
            mapping_config = {
                'table_name': 'retail_data',
                'fields': {self.mapping_table.item(row, 0).text(): 
                          self.mapping_table.item(row, 1).text()
                          for row in range(self.mapping_table.rowCount())}
            }
            
            with open('config.json', 'r+', encoding='utf-8') as f:
                config = json.load(f)
                config['table_mapping'] = mapping_config
                f.seek(0)
                json.dump(config, f, indent=4, ensure_ascii=False)
                f.truncate()
        except Exception as e:
            print(f"保存默认配置失败: {str(e)}")
    
    def initUI(self):
        layout = QVBoxLayout()
        
        # 历史配置选择组
        history_group = QGroupBox("历史映射配置")
        history_layout = QVBoxLayout()
        
        # 添加历史配置下拉框
        self.history_combo = QComboBox()
        self.history_combo.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 4px;
                min-width: 200px;
            }
        """)
        self.history_combo.currentIndexChanged.connect(self.on_history_selected)
        
        # 添加历史配置操作按钮
        history_button_layout = QHBoxLayout()
        
        save_as_button = QPushButton("保存为新配置")
        save_as_button.clicked.connect(self.save_as_new_config)
        save_as_button.setStyleSheet(self.get_button_style('#2196F3'))
        
        delete_config_button = QPushButton("删除当前配置")
        delete_config_button.clicked.connect(self.delete_current_config)
        delete_config_button.setStyleSheet(self.get_button_style('#f44336'))
        
        history_button_layout.addWidget(save_as_button)
        history_button_layout.addWidget(delete_config_button)
        
        history_layout.addWidget(self.history_combo)
        history_layout.addLayout(history_button_layout)
        history_group.setLayout(history_layout)
        layout.addWidget(history_group)
        
        # 表名配置
        table_group = QGroupBox("数据表配置")
        table_layout = QFormLayout()
        
        self.table_name = QLineEdit()
        table_layout.addRow("表名:", self.table_name)
        
        table_group.setLayout(table_layout)
        layout.addWidget(table_group)
        
        # 字段映射配置
        mapping_group = QGroupBox("字段映射配置")
        mapping_layout = QVBoxLayout()
        
        # 创建分割器容器
        splitter_container = QWidget()
        splitter_layout = QVBoxLayout(splitter_container)
        splitter_layout.setContentsMargins(0, 0, 0, 0)
        
        # 创建表格
        self.mapping_table = QTableWidget()
        self.mapping_table.setColumnCount(3)
        self.mapping_table.setHorizontalHeaderLabels(['数据库字段', 'API字段', '字段说明'])
        self.mapping_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # 设置表格最小和最大高度
        self.mapping_table.setMinimumHeight(200)
        self.mapping_table.setMaximumHeight(600)
        
        # 添加表格大小调整功能
        size_grip = QSizeGrip(self.mapping_table)
        size_grip.setStyleSheet("""
            QSizeGrip {
                background-color: #e0e0e0;
                width: 16px;
                height: 16px;
                margin: 2px;
            }
            QSizeGrip:hover {
                background-color: #bdbdbd;
            }
        """)
        
        # 创建一个包含表格和大小调整手柄的容器
        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.addWidget(self.mapping_table)
        
        # 创建����柄��器
        grip_container = QWidget()
        grip_layout = QHBoxLayout(grip_container)
        grip_layout.setContentsMargins(0, 0, 0, 0)
        grip_layout.addStretch()
        grip_layout.addWidget(size_grip)
        
        table_layout.addWidget(grip_container)
        
        # 设置表格样式
        self.mapping_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #d0d0d0;
                border-radius: 4px;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #e3f2fd;
                color: black;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 5px;
                border: none;
                border-right: 1px solid #d0d0d0;
                border-bottom: 1px solid #d0d0d0;
            }
            QScrollBar:vertical {
                border: none;
                background: #f0f0f0;
                width: 10px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #c0c0c0;
                min-height: 20px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical:hover {
                background: #a0a0a0;
            }
        """)
        
        splitter_layout.addWidget(table_container)
        mapping_layout.addWidget(splitter_container)
        
        # 添加说明标签
        help_label = QLabel("""
        字段说明：
        - 数据库字段：您的数据库中实际的字段名
        - API字段：系统要求的标准字段名
        - 字段说明：字段的说明
        
        注意：
        1. 字段映射必须一一对应
        2. API字段名称必须与系统要求一致
        3. 数据库字段名可以根据实际情况修改
        """)
        help_label.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                padding: 10px;
                border-radius: 4px;
                border: 1px solid #dee2e6;
                color: #495057;
            }
        """)
        mapping_layout.addWidget(help_label)
        
        # 添加/删除映射按钮
        button_layout = QHBoxLayout()
        
        add_button = QPushButton("添加映射")
        add_button.clicked.connect(self.add_mapping)
        add_button.setStyleSheet('''
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        ''')
        
        delete_button = QPushButton("删除映射")
        delete_button.clicked.connect(self.delete_mapping)
        delete_button.setStyleSheet('''
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        ''')
        
        view_default_button = QPushButton("查看默认映射")
        view_default_button.clicked.connect(self.view_default_mapping)
        view_default_button.setStyleSheet('''
            QPushButton {
                background-color: #673AB7;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #5E35B1;
            }
        ''')
        
        button_layout.addWidget(add_button)
        button_layout.addWidget(delete_button)
        button_layout.addWidget(view_default_button)
        
        mapping_layout.addLayout(button_layout)
        mapping_group.setLayout(mapping_layout)
        layout.addWidget(mapping_group)
        
        # 保存和重置按钮
        action_layout = QHBoxLayout()
        
        save_button = QPushButton("保存映射配置")
        save_button.clicked.connect(self.save_mapping)
        save_button.setStyleSheet('''
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        ''')
        
        reset_button = QPushButton("重置为默认")
        reset_button.clicked.connect(self.reset_to_default)
        reset_button.setStyleSheet('''
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
        ''')
        
        action_layout.addWidget(save_button)
        action_layout.addWidget(reset_button)
        
        layout.addLayout(action_layout)
        
        self.setLayout(layout)
        
        # 更新历史配置下拉框
        self.update_history_combo()
        
    def get_button_style(self, color):
        """获取按钮样式"""
        return f'''
            QPushButton {{
                background-color: {color};
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }}
            QPushButton:hover {{
                background-color: {color}dd;
            }}
        '''
        
    def load_mapping_history(self):
        """加载历史映射配置"""
        try:
            if os.path.exists(self.mapping_history_file):
                with open(self.mapping_history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return {'configurations': []}
        except Exception:
            return {'configurations': []}
            
    def save_mapping_history(self):
        """保存历史映射配置"""
        try:
            with open(self.mapping_history_file, 'w', encoding='utf-8') as f:
                json.dump(self.mapping_history, f, indent=4, ensure_ascii=False)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存历史配置失败: {str(e)}")
            
    def update_history_combo(self):
        """更新历史配置下拉框"""
        self.history_combo.clear()
        self.history_combo.addItem("默认配置")
        for config in self.mapping_history['configurations']:
            self.history_combo.addItem(config['name'])
            
    def on_history_selected(self, index):
        """历史配置选择改变时"""
        try:
            if index == 0:  # 默认配置
                self.reset_to_default()
                # 保存默认配置到 config.json
                self.save_mapping()
                return
                
            config = self.mapping_history['configurations'][index - 1]
            
            # 验证表名
            table_name = config.get('table_name', '').strip()
            if not table_name or not table_name.replace('_', '').isalnum():
                QMessageBox.warning(self, "错误", f"配置 '{config.get('name')}' 的表名无效！")
                self.reset_to_default()
                return
                
            self.table_name.setText(table_name)
            
            # 清空当前表格
            self.mapping_table.setRowCount(0)
            
            # 填充新的映射关系
            fields = config.get('fields', {})
            if not fields:
                QMessageBox.warning(self, "错误", f"配置 '{config.get('name')}' 没有字段映射！")
                self.reset_to_default()
                return
                
            for i, (db_field, api_field) in enumerate(fields.items()):
                self.mapping_table.insertRow(i)
                self.mapping_table.setItem(i, 0, QTableWidgetItem(db_field))
                self.mapping_table.setItem(i, 1, QTableWidgetItem(api_field))
                
                # 查找并添加字段说明
                description = ""
                for _, default_api, desc in self.default_mappings:
                    if default_api == api_field:
                        description = desc
                        break
                self.mapping_table.setItem(i, 2, QTableWidgetItem(description))
                
            # 立即保存到 config.json
            self.save_mapping()
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载配置失败: {str(e)}")
            self.reset_to_default()

    def save_as_new_config(self):
        """保存为新配置"""
        name, ok = QInputDialog.getText(self, '保存配置', '请输入配置名称:')
        if ok and name:
            # 检查是否存在同名配置
            for config in self.mapping_history['configurations']:
                if config['name'] == name:
                    QMessageBox.warning(self, "错误", "已存在同名配置！")
                    return
                    
            # 收集当前配置
            current_config = {
                'name': name,
                'table_name': self.table_name.text(),
                'fields': {}
            }
            
            for row in range(self.mapping_table.rowCount()):
                db_field = self.mapping_table.item(row, 0)
                api_field = self.mapping_table.item(row, 1)
                if db_field and api_field:
                    current_config['fields'][db_field.text()] = api_field.text()
                    
            # 添加到历史记录
            self.mapping_history['configurations'].append(current_config)
            self.save_mapping_history()
            
            # 更新下拉框
            self.update_history_combo()
            
            QMessageBox.information(self, "成功", "配置保存成功！")
            
    def delete_current_config(self):
        """删除当前配置"""
        current_index = self.history_combo.currentIndex()
        if current_index == 0:
            QMessageBox.warning(self, "错误", "不能删除默认配置！")
            return
            
        reply = QMessageBox.question(self, '确认删除', 
                                   "确定要删除当前配置吗？此操作不可复。",
                                   QMessageBox.Yes | QMessageBox.No, 
                                   QMessageBox.No)
                                   
        if reply == QMessageBox.Yes:
            self.mapping_history['configurations'].pop(current_index - 1)
            self.save_mapping_history()
            self.update_history_combo()
            QMessageBox.information(self, "成功", "配置已删除！")
        
    def reset_to_default(self):
        """重置为默认映射配置"""
        # 设置默认表名
        self.table_name.setText('retail_data')
        
        # 清空当前表格
        self.mapping_table.setRowCount(0)
        
        # 添加默认映射
        for i, (db_field, api_field, description) in enumerate(self.default_mappings):
            self.mapping_table.insertRow(i)
            self.mapping_table.setItem(i, 0, QTableWidgetItem(db_field))
            self.mapping_table.setItem(i, 1, QTableWidgetItem(api_field))
            self.mapping_table.setItem(i, 2, QTableWidgetItem(description))
        
        # 静默保存默认配置
        try:
            mapping_config = {
                'table_name': 'retail_data',
                'fields': {self.mapping_table.item(row, 0).text(): 
                          self.mapping_table.item(row, 1).text()
                          for row in range(self.mapping_table.rowCount())}
            }
            
            with open('config.json', 'r+', encoding='utf-8') as f:
                config = json.load(f)
                config['table_mapping'] = mapping_config
                f.seek(0)
                json.dump(config, f, indent=4, ensure_ascii=False)
                f.truncate()
        except Exception as e:
            print(f"保存默认配置失败: {str(e)}")

    def add_mapping(self):
        """添加新的映射行"""
        dialog = QDialog(self)
        dialog.setWindowTitle("添加新映射")
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout()
        
        # 添加说明
        help_label = QLabel("""
        请选择要映射的API字段，然后输入您的数据库字段名。
        注意：API字段名称不可修改，必须与系统要求一致。
        """)
        help_label.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                padding: 10px;
                border-radius: 4px;
                border: 1px solid #dee2e6;
                color: #495057;
            }
        """)
        layout.addWidget(help_label)
        
        # 创建表单
        form_layout = QFormLayout()
        
        # API字段选择下拉框
        api_field_combo = QComboBox()
        api_field_combo.setStyleSheet("""
            QComboBox {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 3px;
                min-width: 200px;
            }
        """)
        
        # 添加所有可用的API字段
        for _, api_field, description in self.default_mappings:
            api_field_combo.addItem(f"{api_field} ({description})", api_field)
            
        # 数据库字段输入框
        db_field_input = QLineEdit()
        db_field_input.setPlaceholderText("输入您的数据库字段名")
        db_field_input.setStyleSheet("""
            QLineEdit {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 3px;
                min-width: 200px;
            }
        """)
        
        form_layout.addRow("选择API字段:", api_field_combo)
        form_layout.addRow("数据库字段名:", db_field_input)
        
        layout.addLayout(form_layout)
        
        # 按钮布局
        button_layout = QHBoxLayout()
        
        add_button = QPushButton("添加")
        add_button.setStyleSheet('''
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        ''')
        
        cancel_button = QPushButton("取消")
        cancel_button.setStyleSheet('''
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        ''')
        
        button_layout.addWidget(add_button)
        button_layout.addWidget(cancel_button)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        
        dialog.setLayout(layout)
        
        # 连���按钮信号
        def on_add():
            api_field = api_field_combo.currentData()
            db_field = db_field_input.text().strip()
            
            if not db_field:
                QMessageBox.warning(dialog, "错误", "请输入数据库字段��！")
                return
                
            # 检查是否已存在相同的映射
            for row in range(self.mapping_table.rowCount()):
                existing_api = self.mapping_table.item(row, 1)
                if existing_api and existing_api.text() == api_field:
                    QMessageBox.warning(dialog, "错误", "该API字段已经存在映射关系！")
                    return
                    
            # 添加新行
            row = self.mapping_table.rowCount()
            self.mapping_table.insertRow(row)
            
            # 获取选中API字段的描述
            description = ""
            for _, api, desc in self.default_mappings:
                if api == api_field:
                    description = desc
                    break
                    
            self.mapping_table.setItem(row, 0, QTableWidgetItem(db_field))
            self.mapping_table.setItem(row, 1, QTableWidgetItem(api_field))
            self.mapping_table.setItem(row, 2, QTableWidgetItem(description))
            
            dialog.accept()
            
        add_button.clicked.connect(on_add)
        cancel_button.clicked.connect(dialog.reject)
        
        dialog.exec_()

    def delete_mapping(self):
        """删除��中的映射行"""
        current_row = self.mapping_table.currentRow()
        if current_row >= 0:
            self.mapping_table.removeRow(current_row)
            
    def save_mapping(self):
        """保映射配置"""
        # 验证表名
        table_name = self.table_name.text().strip()
        if not table_name:
            QMessageBox.warning(self, "错误", "表名不能为空！")
            return
            
        # 验证表名格式（只允许字母数字、下划线）
        if not table_name.replace('_', '').isalnum():
            QMessageBox.warning(self, "错误", "表名只能包含字母、数字和下划线！")
            return
            
        mapping_config = {
            'table_name': table_name,
            'fields': {}
        }
        
        # 验证是否有映射关系
        if self.mapping_table.rowCount() == 0:
            QMessageBox.warning(self, "错误", "请至少添加一个字段映射！")
            return
            
        for row in range(self.mapping_table.rowCount()):
            db_field = self.mapping_table.item(row, 0)
            api_field = self.mapping_table.item(row, 1)
            if db_field and api_field:
                mapping_config['fields'][db_field.text()] = api_field.text()
                
        try:
            with open('config.json', 'r+', encoding='utf-8') as f:
                config = json.load(f)
                config['table_mapping'] = mapping_config
                f.seek(0)
                json.dump(config, f, indent=4, ensure_ascii=False)
                f.truncate()
            QMessageBox.information(self, "成功", "���射配���保存成功！")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存映射配置失败: {str(e)}")
            
    def view_default_mapping(self):
        """查看默认映射关系"""
        dialog = QDialog(self)
        dialog.setWindowTitle("默认字段映射关系")
        dialog.setMinimumWidth(600)
        
        layout = QVBoxLayout()
        
        # 创建表格
        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(['数据库字段', 'API字段', '字段说明'])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # 填充默认映射数据
        table.setRowCount(len(self.default_mappings))
        for i, (db_field, api_field, description) in enumerate(self.default_mappings):
            table.setItem(i, 0, QTableWidgetItem(db_field))
            table.setItem(i, 1, QTableWidgetItem(api_field))
            table.setItem(i, 2, QTableWidgetItem(description))
            
        # 设置表格只读
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        layout.addWidget(table)
        
        # 添加关闭按钮
        close_button = QPushButton("关闭")
        close_button.clicked.connect(dialog.close)
        close_button.setStyleSheet('''
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        ''')
        
        layout.addWidget(close_button, alignment=Qt.AlignRight)
        
        dialog.setLayout(layout)
        dialog.exec_()

    def resizeEvent(self, event):
        """处理窗口大小改事件"""
        super().resizeEvent(event)
        # 更新表格的最大高度为窗口高度的70%
        self.mapping_table.setMaximumHeight(int(self.height() * 0.7))

class ImportTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = None  # 添加主窗口引用
        self.initUI()
    
    def set_main_window(self, main_window):
        """设置主窗口引用"""
        self.main_window = main_window
    
    def initUI(self):
        layout = QVBoxLayout()
        
        # 添加模板生成按钮
        template_button = QPushButton("生成导入模板")
        template_button.clicked.connect(self.create_template)
        template_button.setStyleSheet('''
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 120px;
                margin-bottom: 10px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        ''')
        layout.addWidget(template_button, alignment=Qt.AlignRight)
        
        # 文件选择区域
        file_group = QGroupBox("Excel文件导入")
        file_layout = QVBoxLayout()
        
        # 文件路径显示和选择按钮
        file_select_layout = QHBoxLayout()
        self.file_path = QLineEdit()
        self.file_path.setReadOnly(True)
        self.file_path.setPlaceholderText("请选择Excel文件...")
        
        select_button = QPushButton("选择文件")
        select_button.clicked.connect(self.select_file)
        select_button.setStyleSheet('''
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        ''')
        
        file_select_layout.addWidget(self.file_path)
        file_select_layout.addWidget(select_button)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        
        # 预览表格
        self.preview_table = QTableWidget()
        self.preview_table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        file_layout.addLayout(file_select_layout)
        file_layout.addWidget(self.progress_bar)
        file_layout.addWidget(self.preview_table)
        
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)
        
        # 导入控制按钮
        button_layout = QHBoxLayout()
        
        self.import_button = QPushButton("导入数据")
        self.import_button.clicked.connect(self.import_data)
        self.import_button.setEnabled(False)
        self.import_button.setStyleSheet('''
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        ''')
        
        self.upload_button = QPushButton("上报数据")
        self.upload_button.clicked.connect(self.upload_data)
        self.upload_button.setEnabled(False)
        self.upload_button.setStyleSheet('''
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        ''')
        
        button_layout.addWidget(self.import_button)
        button_layout.addWidget(self.upload_button)
        
        layout.addLayout(button_layout)
        
        # 添加说明文本
        help_text = QLabel("""
        Excel文件格式说明：
        1. 文件必须包含以下字段：
           - 统一社会信用代码
           - 企业名称
           - 零售点编码
           等必要字段
        2. 表头名称必须与系统字段对应
        3. 日期格式：YYYY-MM-DD
        4. 数值字段必须为数字
        """)
        help_text.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                padding: 10px;
                border-radius: 4px;
                border: 1px solid #dee2e6;
                color: #495057;
            }
        """)
        layout.addWidget(help_text)
        
        # 添加映射配置选择
        mapping_layout = QHBoxLayout()
        self.mapping_combo = QComboBox()
        self.mapping_combo.addItem("默认映射")
        self.load_excel_mappings()
        
        mapping_layout.addWidget(QLabel("选择映射配置:"))
        mapping_layout.addWidget(self.mapping_combo)
        layout.addLayout(mapping_layout)
        
        self.setLayout(layout)
        
        # 存储导入的数据
        self.imported_data = None
        
    def select_file(self):
        """选择Excel文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择Excel文件",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.file_path.setText(file_path)
            self.import_button.setEnabled(True)
            
    def load_excel_mappings(self):
        """加载Excel映射配置"""
        try:
            with open('excel_mapping_history.json', 'r', encoding='utf-8') as f:
                mappings = json.load(f)
                for config in mappings['configurations']:
                    self.mapping_combo.addItem(config['name'])
        except:
            pass
            
    def import_data(self):
        """导入Excel数据"""
        try:
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            
            # 读取Excel文件
            df = pd.read_excel(self.file_path.text())
            
            # 处理列名，移除API字段名提示
            df.columns = df.columns.map(lambda x: x.split(' (')[0] if ' (' in x else x)
            
            # 打印原始列名用于调试
            print("原始列名:", df.columns.tolist())
            
            # 获取选择的映射配置
            mapping_name = self.mapping_combo.currentText()
            
            # 设置默认映射关系
            field_mapping = {
                '统一社会信用代码': 'socialCreditCode',
                '企业名称': 'compName',
                '零售点编码': 'retailStoreCode',
                '零售点名称': 'retailStoreName',
                '上报日期': 'reportDate',
                '商品编码': 'selfCommondityCode',
                '商品名称': 'selfCommondityName',
                '单位': 'unit',
                '规格': 'spec',
                '条码': 'barcode',
                '数据类型': 'dataType',
                '数据值': 'dataValue',
                '转换标志': 'dataConvertFlag',
                '供应商编码': 'supplierCode',
                '供应商名称': 'supplierName',
                '生产商名称': 'manufatureName',
                '产地编码': 'originCode',
                '产地名称': 'originName',
                '场景标志': 'sceneflag'
            }
            
            # 如果不是默认映射，则加载自定义映射配置
            if mapping_name != "默认映射":
                try:
                    with open('excel_mapping_history.json', 'r', encoding='utf-8') as f:
                        mappings = json.load(f)
                        for config in mappings['configurations']:
                            if config['name'] == mapping_name:
                                field_mapping = config['mappings']
                                break
                except Exception as e:
                    print(f"加载自定义映射配置失败: {str(e)}")
                    # 如果加载失败，继续使用默认映射
            
            # 使用映射配置重命名列
            df = df.rename(columns=field_mapping)
            
            # 打印映射后的名用于调试
            print("映射后的列名:", df.columns.tolist())
            
            # 验证必要字段
            required_fields = {
                'socialCreditCode': '统一社会信用代码',
                'compName': '企业名称',
                'retailStoreCode': '零售点编码',
                'retailStoreName': '零售点名称',
                'reportDate': '上报日期',
                'selfCommondityCode': '商品编码',
                'selfCommondityName': '商品名称',
                'dataType': '数据类型',
                'dataValue': '数据值',
                'dataConvertFlag': '转换标志',
                'supplierCode': '供应商编码',
                'supplierName': '供应商名称',
                'manufatureName': '生产商名称',
                'originCode': '产地编码',
                'originName': '产地名称',
                'sceneflag': '场��标志'
            }
            
            # 检查字段是否存在
            missing_fields = []
            for field, field_name in required_fields.items():
                if field not in df.columns:
                    missing_fields.append(field_name)
                    print(f"缺少字段: {field_name} ({field})")  # 调试信息
                    
            if missing_fields:
                QMessageBox.warning(self, "错误", f"缺少必要字段：{', '.join(missing_fields)}")
                return
                
            # 数据类型转换和验证
            try:
                # 转换日期格式
                df['reportDate'] = pd.to_datetime(df['reportDate']).dt.strftime('%Y-%m-%d')
                
                # 确保数值字段为数字类型
                df['dataType'] = df['dataType'].astype(int)
                df['dataValue'] = df['dataValue'].astype(float)
                df['dataConvertFlag'] = df['dataConvertFlag'].astype(int)
                df['sceneflag'] = df['sceneflag'].astype(int)
                
                # 生成itemId
                df['itemId'] = 'YN' + df['reportDate'].str.replace('-', '') + df.index.astype(str).str.zfill(6)
                
            except Exception as e:
                QMessageBox.warning(self, "错误", f"数据格式转换失败: {str(e)}")
                return
                
            # 更新进度条
            self.progress_bar.setValue(30)
            
            # 显示预览
            self.preview_table.setRowCount(len(df))
            self.preview_table.setColumnCount(len(df.columns))
            self.preview_table.setHorizontalHeaderLabels(df.columns)
            
            for i in range(len(df)):
                for j in range(len(df.columns)):
                    value = str(df.iloc[i, j])
                    self.preview_table.setItem(i, j, QTableWidgetItem(value))
                    
            # 存储导入的数据
            self.imported_data = df.to_dict('records')
            
            # 更新进度条
            self.progress_bar.setValue(100)
            
            # 启用上报按钮
            self.upload_button.setEnabled(True)
            
            QMessageBox.information(self, "成功", f"成功导入 {len(df)} 条数据！")
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导入失败: {str(e)}")
        finally:
            self.progress_bar.setVisible(False)
            
    def upload_data(self):
        """上报导入的数据"""
        if not self.imported_data:
            QMessageBox.warning(self, "错误", "没有可上报的数据！")
            return
            
        reply = QMessageBox.question(
            self,
            "确认",
            f"确定要上报 {len(self.imported_data)} 条数据吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                # 初始化API客户端
                with open('config.json', 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    
                api = RetailAPI(config['api']['url'])
                if not api.login(config['api']['username'], config['api']['password']):
                    QMessageBox.warning(self, "错误", "API登录失败！")
                    return
                    
                # 上报数据
                result = api.upload_retail_data(self.imported_data)
                
                if result and result.get("code") == 200:
                    QMessageBox.information(self, "成功", "数据上报成功！")
                    # 保存成功历史
                    self.save_history(
                        status='成功',
                        data_count=len(self.imported_data),
                        message=str(result.get("content", [])),
                        error_detail=None,
                        source='Excel导入'
                    )
                    # 刷新历史记录
                    if self.main_window:
                        self.main_window.refresh_history()
                else:
                    QMessageBox.warning(self, "错误", f"上报失败: {str(result)}")
                    # 保存失败历史
                    self.save_history(
                        status='失败',
                        data_count=len(self.imported_data),
                        message="上报失败",
                        error_detail=str(result),
                        source='Excel导入'  # 添加数据来源标识
                    )
                    
            except Exception as e:
                QMessageBox.warning(self, "错误", f"上报失败: {str(e)}")
                # 保存错误历史
                self.save_history(
                    status='失败',
                    data_count=0,
                    message="执行出错",
                    error_detail=str(e),
                    source='Excel导入'  # 添加数据来源标识
                )
            
    def save_history(self, status, data_count, message, error_detail=None, source='Excel导入'):
        """保存上报历史到JSON文件"""
        try:
            history_file = 'upload_history.json'
            history_data = []
            
            # 读取现有历史记录
            if os.path.exists(history_file):
                try:
                    with open(history_file, 'r', encoding='utf-8') as f:
                        history_data = json.load(f)
                except:
                    history_data = []
            
            # 添加新记录
            new_record = {
                'upload_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'status': status,
                'data_count': data_count,
                'message': message,
                'error_detail': error_detail,
                'source': source  # 添加数据来源标识
            }
            
            # 将新记录添加到开头
            history_data.insert(0, new_record)
            
            # 只保留最近100条记录
            history_data = history_data[:100]
            
            # 保存到文件
            with open(history_file, 'w', encoding='utf-8') as f:
                json.dump(history_data, f, indent=4, ensure_ascii=False)
                
        except Exception as e:
            print(f"保存历史记录失败: {str(e)}")

    def create_template(self):
        """生成Excel模板文件"""
        try:
            # 选择保存位置
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "保存模板文件",
                "数据导入模板.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
                
            import pandas as pd
            import numpy as np
            from datetime import datetime
            from openpyxl.styles import PatternFill, Font
            
            # 创建示例数据
            example_data = {
                '统一��会信用代码 (socialCreditCode)': ['91532901792864164X1'] * 4,
                '企业名称 (compName)': ['云南市方街商贸有限公司'] * 4,
                '零售点编码 (retailStoreCode)': ['SFJRPA1234'] * 4,
                '零售点名称 (retailStoreName)': ['四方街商贸零售点'] * 4,
                '上报日期 (reportDate)': [datetime.now().strftime('%Y-%m-%d')] * 4,
                '商品编码 (selfCommondityCode)': ['170060'] * 4,
                '商品名称 (selfCommondityName)': ['大白菜'] * 4,
                '单位 (unit)': ['公斤'] * 4,
                '规格 (spec)': ['散装'] * 4,
                '条码 (barcode)': ['170060'] * 4,
                '数据类型 (dataType)': [1, 2, 3, 4],  # 初始库存、入库量、销售量、价格
                '数据值 (dataValue)': [100, 80, 50, 7.00],
                '转换标志 (dataConvertFlag)': [2] * 4,
                '供应商编码 (supplierCode)': ['SUP001'] * 4,
                '供应商名称 (supplierName)': ['大理批发市场'] * 4,
                '生产商名称 (manufatureName)': ['大理蔬菜基地'] * 4,
                '产地编码 (originCode)': ['530000'] * 4,
                '产地名称 (originName)': ['云南省'] * 4,
                '场景标志 (sceneflag)': [1] * 4
            }
            
            # 建DataFrame
            df = pd.DataFrame(example_data)
            
            # 添加说明行
            description_data = {
                '统一社会信用代码 (socialCreditCode)': '企业统一社会信用代码',
                '企业名称 (compName)': '企业全称',
                '零售点编码 (retailStoreCode)': '零售点唯一编码',
                '零售点名称 (retailStoreName)': '零售点名称',
                '上报日期 (reportDate)': '数据日期（格式：YYYY-MM-DD）',
                '商品编码 (selfCommondityCode)': '商品唯一编码',
                '商品名称 (selfCommondityName)': '商品名称',
                '单位 (unit)': '计量单位',
                '规格 (spec)': '商品规格',
                '条码 (barcode)': '商品条形码',
                '数据类型 (dataType)': '1期初库存、2入库量、3销售量、4价格',
                '数据值 (dataValue)': '对应数据类型的数值',
                '转换标志 (dataConvertFlag)': '默认值2',
                '供应商编码 (supplierCode)': '供应商编码',
                '供应商名称 (supplierName)': '供应商名称',
                '生产商名称 (manufatureName)': '生产厂家名称',
                '产地编码 (originCode)': '产地编码（示例：530000）',
                '产地名称 (originName)': '产地名称（示例：云南省）',
                '场景标志 (sceneflag)': '场景标志（默认值1）'
            }
            
            description_df = pd.DataFrame([description_data])
            
            # 合并说明和示例数据
            final_df = pd.concat([description_df, df], ignore_index=True)
            
            # 创建Excel写入器
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 写入数据页
                final_df.to_excel(writer, sheet_name='数据模板', index=False)
                
                # 获取工作簿和工作表
                workbook = writer.book
                worksheet = writer.sheets['数据模板']
                
                # 设置列宽
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                    
                # 设置说明行样式
                yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                red_font = Font(color='FF0000')
                
                for cell in worksheet[1]:
                    cell.fill = yellow_fill
                    cell.font = red_font
                    
            QMessageBox.information(self, "成功", f"模板文件已保存到：\n{file_path}")
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"生成模板文件失败: {str(e)}")

class ExcelMappingTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.mapping_history_file = 'excel_mapping_history.json'
        self.mapping_history = self.load_mapping_history()
        self.initUI()
        
    def initUI(self):
        """初始化UI"""
        layout = QVBoxLayout()
        
        # 添加默认模板展示按钮
        template_button = QPushButton("查看默认模板")
        template_button.clicked.connect(self.show_default_template)
        template_button.setStyleSheet('''
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 120px;
                margin-bottom: 10px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        ''')
        layout.addWidget(template_button, alignment=Qt.AlignRight)
        
        # 配置选择组
        config_group = QGroupBox("Excel映射配置")
        config_layout = QVBoxLayout()
        
        # 配置名称选择
        config_select_layout = QHBoxLayout()
        self.config_combo = QComboBox()
        self.config_combo.currentIndexChanged.connect(self.on_config_selected)
        
        new_config_button = QPushButton("新建配置")
        new_config_button.clicked.connect(self.create_new_config)
        delete_config_button = QPushButton("删除配置")
        delete_config_button.clicked.connect(self.delete_config)
        
        config_select_layout.addWidget(QLabel("选择配置:"))
        config_select_layout.addWidget(self.config_combo)
        config_select_layout.addWidget(new_config_button)
        config_select_layout.addWidget(delete_config_button)
        
        config_layout.addLayout(config_select_layout)
        config_group.setLayout(config_layout)
        layout.addWidget(config_group)
        
        # 映射表格
        mapping_group = QGroupBox("字段映射关系")
        mapping_layout = QVBoxLayout()
        
        self.mapping_table = QTableWidget()
        self.mapping_table.setColumnCount(3)
        self.mapping_table.setHorizontalHeaderLabels(['Excel表头', '系统字段', '说明'])
        self.mapping_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # 添加/删除映射按钮
        button_layout = QHBoxLayout()
        add_button = QPushButton("添加映射")
        add_button.clicked.connect(self.add_mapping)
        delete_button = QPushButton("删除映射")
        delete_button.clicked.connect(self.delete_mapping)
        
        button_layout.addWidget(add_button)
        button_layout.addWidget(delete_button)
        
        mapping_layout.addWidget(self.mapping_table)
        mapping_layout.addLayout(button_layout)
        
        # 保存按钮
        save_button = QPushButton("保存配置")
        save_button.clicked.connect(self.save_config)
        mapping_layout.addWidget(save_button)
        
        mapping_group.setLayout(mapping_layout)
        layout.addWidget(mapping_group)
        
        # 帮助说明
        help_text = QLabel("""
        使用说明���
        1. 创建新的映射配置并命名
        2. 添加Excel表格中的表头与系统字段的对应关系
        3. 保存配置后即可在导入Excel时使用
        
        注意：
        - Excel表头填写实际的表格中的列名
        - 系统字段必须与标准字段完全匹��
        - 必须包含所有必要字段的映射关系
        """)
        help_text.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                padding: 10px;
                border-radius: 4px;
                border: 1px solid #dee2e6;
                color: #495057;
            }
        """)
        layout.addWidget(help_text)
        
        self.setLayout(layout)
        self.update_config_list()
        
    def create_new_config(self):
        """创建新配置"""
        dialog = QDialog(self)
        dialog.setWindowTitle("新建映射配置")
        dialog.setMinimumWidth(800)
        dialog.setMinimumHeight(600)
        
        layout = QVBoxLayout()
        
        # 配置名称输入
        name_layout = QHBoxLayout()
        name_layout.addWidget(QLabel("配置名称:"))
        name_input = QLineEdit()
        name_layout.addWidget(name_input)
        
        # 添加导入按钮
        import_button = QPushButton("导入Excel表头")
        import_button.setStyleSheet('''
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        ''')
        name_layout.addWidget(import_button)
        
        layout.addLayout(name_layout)
        
        # 创建映射表格
        mapping_table = QTableWidget()
        mapping_table.setColumnCount(4)
        mapping_table.setHorizontalHeaderLabels(['Excel表头', '系统字段', '说明', '是否映射'])
        mapping_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(mapping_table)
        
        # 获取系统字段列表
        system_fields = self.get_system_fields()
        
        def import_excel():
            """导入Excel获取表头"""
            file_path, _ = QFileDialog.getOpenFileName(
                dialog,
                "选择Excel文件",
                "",
                "Excel Files (*.xlsx *.xls)"
            )
            
            if file_path:
                try:
                    # 读取Excel表头
                    df = pd.read_excel(file_path, nrows=0)
                    excel_headers = df.columns.tolist()
                    
                    # 清空表格
                    mapping_table.setRowCount(0)
                    
                    # 添加所有Excel表头
                    for header in excel_headers:
                        row = mapping_table.rowCount()
                        mapping_table.insertRow(row)
                        
                        # Excel表头
                        mapping_table.setItem(row, 0, QTableWidgetItem(header))
                        
                        # 添加系统字段下拉框
                        combo = QComboBox()
                        combo.addItem("-- 请选择 --", "")
                        for field, desc in system_fields:
                            combo.addItem(f"{field} ({desc})", field)
                        mapping_table.setCellWidget(row, 1, combo)
                        
                        # 添加说明列
                        mapping_table.setItem(row, 2, QTableWidgetItem(""))
                        
                        # 添加复选框
                        checkbox = QCheckBox()
                        checkbox_widget = QWidget()
                        checkbox_layout = QHBoxLayout(checkbox_widget)
                        checkbox_layout.addWidget(checkbox)
                        checkbox_layout.setAlignment(Qt.AlignCenter)
                        checkbox_layout.setContentsMargins(0, 0, 0, 0)
                        mapping_table.setCellWidget(row, 3, checkbox_widget)
                        
                        # 自动匹配系统字段
                        for i in range(combo.count()):
                            field_data = combo.itemData(i)
                            if field_data and field_data.lower() in header.lower().replace(" ", ""):
                                combo.setCurrentIndex(i)
                                checkbox.setChecked(True)
                                # 更新说明
                                for field, desc in system_fields:
                                    if field == field_data:
                                        mapping_table.setItem(row, 2, QTableWidgetItem(desc))
                                        break
                                break
                        
                        # 连接下拉框信号
                        def on_field_selected(index, row=row):
                            field = combo.itemData(index)
                            checkbox = mapping_table.cellWidget(row, 3).findChild(QCheckBox)
                            checkbox.setChecked(bool(field))
                            # 更新说明
                            for f, desc in system_fields:
                                if f == field:
                                    mapping_table.setItem(row, 2, QTableWidgetItem(desc))
                                    break
                        
                        combo.currentIndexChanged.connect(on_field_selected)
                        
                except Exception as e:
                    QMessageBox.warning(dialog, "错误", f"导入Excel失败: {str(e)}")
        
        import_button.clicked.connect(import_excel)
        
        # 添加按钮
        button_layout = QHBoxLayout()
        save_button = QPushButton("保存")
        cancel_button = QPushButton("取消")
        
        save_button.setStyleSheet('''
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        ''')
        
        cancel_button.setStyleSheet('''
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        ''')
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        layout.addLayout(button_layout)
        
        def save_config():
            """保存配置"""
            name = name_input.text().strip()
            if not name:
                QMessageBox.warning(dialog, "错误", "请输入配置名称！")
                return
            
            if name in [self.config_combo.itemText(i) for i in range(self.config_combo.count())]:
                QMessageBox.warning(dialog, "错误", "配置名称已存在！")
                return
            
            # 收集映射关系
            mappings = {}
            for row in range(mapping_table.rowCount()):
                checkbox = mapping_table.cellWidget(row, 3).findChild(QCheckBox)
                if checkbox.isChecked():
                    excel_field = mapping_table.item(row, 0).text()
                    combo = mapping_table.cellWidget(row, 1)
                    system_field = combo.currentData()
                    if system_field:
                        mappings[excel_field] = system_field
                        
            if not mappings:
                QMessageBox.warning(dialog, "错误", "请至少选择一个映射关系！")
                return
            
            # 保存配置
            self.mapping_history['configurations'].append({
                'name': name,
                'mappings': mappings
            })
            self.save_mapping_history()
            self.update_config_list()
            self.config_combo.setCurrentText(name)
            
            QMessageBox.information(dialog, "成功", "配置保存成功！")
            dialog.accept()
        
        save_button.clicked.connect(save_config)
        cancel_button.clicked.connect(dialog.reject)
        
        dialog.setLayout(layout)
        dialog.exec_()
    
    def load_mapping_history(self):
        """加载映射配置历史"""
        try:
            if os.path.exists(self.mapping_history_file):
                with open(self.mapping_history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return {'configurations': []}
        except Exception as e:
            print(f"加载映射配置失败: {str(e)}")
            return {'configurations': []}
            
    def save_mapping_history(self):
        """保存映射配置历史"""
        try:
            with open(self.mapping_history_file, 'w', encoding='utf-8') as f:
                json.dump(self.mapping_history, f, indent=4, ensure_ascii=False)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存映射配置失败: {str(e)}")
            
    def update_config_list(self):
        """更新配置列表"""
        self.config_combo.clear()
        
        # 添加默认配置
        default_mappings = {
            '统一社会信用代码': 'socialCreditCode',
            '企业名称': 'compName',
            '零售点编码': 'retailStoreCode',
            '零售点名称': 'retailStoreName',
            '上报日期': 'reportDate',
            '商品编码': 'selfCommondityCode',
            '商品名称': 'selfCommondityName',
            '单位': 'unit',
            '规格': 'spec',
            '条码': 'barcode',
            '数据类型': 'dataType',
            '数据值': 'dataValue',
            '转换标志': 'dataConvertFlag',
            '供应商编码': 'supplierCode',
            '供应商名称': 'supplierName',
            '生产商名称': 'manufatureName',
            '产地编码': 'originCode',
            '产地名称': 'originName',
            '场景标志': 'sceneflag'
        }
        
        # 添加默认配置到历史记录中
        if not any(config['name'] == '默认配置' for config in self.mapping_history['configurations']):
            self.mapping_history['configurations'].insert(0, {
                'name': '默认配置',
                'mappings': default_mappings
            })
            self.save_mapping_history()
        
        # 更新下拉框
        for config in self.mapping_history['configurations']:
            self.config_combo.addItem(config['name'])
            
    def on_config_selected(self, index):
        """配置选择改变时的处理"""
        if index < 0:
            return
            
        config_name = self.config_combo.currentText()
        
        # 加载配置
        for config in self.mapping_history['configurations']:
            if config['name'] == config_name:
                self.load_mapping_table(config['mappings'])
                
                # 如果是默认配置，禁用删除按钮
                delete_button = self.findChild(QPushButton, "delete_config_button")
                if delete_button:
                    delete_button.setEnabled(config_name != '默认配置')
                break
                
    def load_mapping_table(self, mappings):
        """加载映射关系到表格"""
        self.mapping_table.setRowCount(0)
        for excel_field, system_field in mappings.items():
            row = self.mapping_table.rowCount()
            self.mapping_table.insertRow(row)
            self.mapping_table.setItem(row, 0, QTableWidgetItem(excel_field))
            self.mapping_table.setItem(row, 1, QTableWidgetItem(system_field))
            # 找到对应的说明
            for field, desc in self.get_system_fields():
                if field == system_field:
                    self.mapping_table.setItem(row, 2, QTableWidgetItem(desc))
                    break
                    
    def get_system_fields(self):
        """获取系统标准字段列表"""
        return [
            ('socialCreditCode', '统一社会信用代码'),
            ('compName', '企业名称'),
            ('retailStoreCode', '零售点编码'),
            ('retailStoreName', '零售点名称'),
            ('reportDate', '上报日期'),
            ('selfCommondityCode', '商品编码'),
            ('selfCommondityName', '商品名称'),
            ('unit', '单位'),
            ('spec', '规格'),
            ('barcode', '条码'),
            ('dataType', '数据类型'),
            ('dataValue', '数据值'),
            ('dataConvertFlag', '转换标志'),
            ('supplierCode', '供应商编码'),
            ('supplierName', '供应商名称'),
            ('manufatureName', '生产商名称'),
            ('originCode', '产地编码'),
            ('originName', '产地名称'),
            ('sceneflag', '场景标志')
        ]
            
    def delete_mapping(self):
        """删除选中的映射关系"""
        current_row = self.mapping_table.currentRow()
        if current_row >= 0:
            self.mapping_table.removeRow(current_row)
            
    def delete_config(self):
        """删除当前配置"""
        current_config = self.config_combo.currentText()
        if not current_config:
            return
            
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定要删除配置 '{current_config}' 吗？此操作不可恢复。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.mapping_history['configurations'] = [
                config for config in self.mapping_history['configurations']
                if config['name'] != current_config
            ]
            self.save_mapping_history()
            self.update_config_list()
            
    def save_config(self):
        """保存当前配置"""
        current_config = self.config_combo.currentText()
        if not current_config:
            QMessageBox.warning(self, "错误", "请先选择或创建一个配置！")
            return
            
        # 收集当前表格中的映射关系
        mappings = {}
        for row in range(self.mapping_table.rowCount()):
            excel_field = self.mapping_table.item(row, 0).text()
            system_field = self.mapping_table.item(row, 1).text()
            mappings[excel_field] = system_field
            
        # 更新配置
        for config in self.mapping_history['configurations']:
            if config['name'] == current_config:
                config['mappings'] = mappings
                break
                
        self.save_mapping_history()
        QMessageBox.information(self, "成功", "配置保存成功！")

    def add_mapping(self):
        """添加新的映射关系"""
        dialog = QDialog(self)
        dialog.setWindowTitle("添加字段映射")
        
        layout = QFormLayout()
        
        excel_field = QLineEdit()
        system_field = QComboBox()
        
        # 添加系统标准字段
        for field, desc in self.get_system_fields():
            system_field.addItem(f"{field} ({desc})", field)
            
        layout.addRow("Excel表头:", excel_field)
        layout.addRow("系统字段:", system_field)
        
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
            Qt.Horizontal, dialog
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        
        layout.addRow(buttons)
        dialog.setLayout(layout)
        
        if dialog.exec_() == QDialog.Accepted:
            if not excel_field.text().strip():
                QMessageBox.warning(self, "错误", "Excel表头不能为空！")
                return
                
            # 检查是否已存在相同的映射
            for row in range(self.mapping_table.rowCount()):
                existing_excel = self.mapping_table.item(row, 0)
                existing_system = self.mapping_table.item(row, 1)
                if existing_excel and existing_excel.text() == excel_field.text():
                    QMessageBox.warning(self, "错误", "该Excel表头已存在映射关系！")
                    return
                if existing_system and existing_system.text() == system_field.currentData():
                    QMessageBox.warning(self, "错误", "该系统字段已存在映射关系！")
                    return
                    
            # 添加新行
            row = self.mapping_table.rowCount()
            self.mapping_table.insertRow(row)
            self.mapping_table.setItem(row, 0, QTableWidgetItem(excel_field.text()))
            selected_field = system_field.currentData()
            self.mapping_table.setItem(row, 1, QTableWidgetItem(selected_field))
            
            # 找到对应的说明
            for field, desc in self.get_system_fields():
                if field == selected_field:
                    self.mapping_table.setItem(row, 2, QTableWidgetItem(desc))
                    break

    def show_default_template(self):
        """展示默认模板"""
        dialog = QDialog(self)
        dialog.setWindowTitle("默认Excel模板")
        dialog.setMinimumWidth(800)
        dialog.setMinimumHeight(600)
        
        layout = QVBoxLayout()
        
        # 创建表格
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(['字段名称', '示例值', '说明', '是否必填'])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # 默认模板数���
        template_data = [
            ('统一社会信用代码', '91532901792864164X1', '企业统一社会信用代码', '是'),
            ('企业名称', '云南市四方街商贸有限公司', '企业全称', '是'),
            ('零售点编码', 'SFJRPA1234', '零售点唯一编码', '是'),
            ('零售点名称', '四方街商贸零售点', '零售点名称', '是'),
            ('上报日期', '2024-03-22', '数据日期（格式：YYYY-MM-DD）', '是'),
            ('商品编码', '170060', '商品唯一编码', '是'),
            ('商品名称', '大白菜', '商品名称', '是'),
            ('单位', '公斤', '计量单位', '是'),
            ('规格', '散装', '商��规格', '是'),
            ('条码', '170060', '商品条形码', '是'),
            ('数据类型', '1/2/3/4', '1期初库存、2入库量、3销售量、4价格', '是'),
            ('数据值', '100.00', '对应数据类型的数值', '是'),
            ('转换标志', '2', '默认值2', '是'),
            ('供应商编码', 'SUP001', '供应商编码', '是'),
            ('供应商名称', '大理批发市场', '供应商名称', '是'),
            ('生产商名称', '大理蔬菜基地', '生产厂家名称', '是'),
            ('产地编码', '530000', '产地编码', '是'),
            ('产地名称', '云南省', '产地名称', '是'),
            ('场景标志', '1', '场景标志（默认值1）', '是')
        ]
        
        # 填充数据
        table.setRowCount(len(template_data))
        for i, (field, example, desc, required) in enumerate(template_data):
            table.setItem(i, 0, QTableWidgetItem(field))
            table.setItem(i, 1, QTableWidgetItem(example))
            table.setItem(i, 2, QTableWidgetItem(desc))
            required_item = QTableWidgetItem(required)
            required_item.setForeground(Qt.red if required == '是' else Qt.black)
            table.setItem(i, 3, required_item)
        
        # ���置表格样式
        table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #d0d0d0;
                border-radius: 4px;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 5px;
                border: none;
                border-right: 1px solid #d0d0d0;
                border-bottom: 1px solid #d0d0d0;
            }
        """)
        
        layout.addWidget(table)
        
        # 添加说明文本
        help_text = QLabel("""
        说明：
        1. 所有字段都必须按照示例格式填写
        2. 日期必须使用 YYYY-MM-DD 格式
        3. 数据类型对应关系：
           - 1: 期初库存
           - 2: 入库量
           - 3: 销售量
           - 4: 价格
        4. 转换标志统一填写 2
        5. 场景标志统一填写 1
        """)
        help_text.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                padding: 10px;
                border-radius: 4px;
                border: 1px solid #dee2e6;
                color: #495057;
            }
        """)
        layout.addWidget(help_text)
        
        # 添加关闭按钮
        close_button = QPushButton("关闭")
        close_button.clicked.connect(dialog.close)
        close_button.setStyleSheet('''
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        ''')
        layout.addWidget(close_button, alignment=Qt.AlignRight)
        
        dialog.setLayout(layout)
        dialog.exec_()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('零售数据上报工具')
        self.setGeometry(100, 100, 1000, 600)
        
        # 创建中央窗口部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        # 创建标签页
        tab_widget = QTabWidget()
        
        # 主页面标签
        main_tab = QWidget()
        main_layout = QVBoxLayout()
        
        # 添加标题标签
        title_label = QLabel('零售数据上报系统')
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet('font-size: 20px; font-weight: bold; margin: 10px;')
        main_layout.addWidget(title_label)
        
        # 添加日志显示区域
        self.log_display = QTextEdit()
        self.log_display.setReadOnly(True)
        main_layout.addWidget(self.log_display)
        
        # 添加按钮
        self.upload_button = QPushButton('开始上报数据')
        self.upload_button.clicked.connect(self.start_upload)
        self.upload_button.setStyleSheet('''
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px;
                border: none;
                border-radius: 5px;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
        ''')
        main_layout.addWidget(self.upload_button)
        
        main_tab.setLayout(main_layout)
        
        # 创建并添加定时任务标签页
        schedule_tab = ScheduleTab()
        schedule_tab.set_main_window(self)  # 设置主窗口引用
        tab_widget.addTab(schedule_tab, "定时任务")
        
        # 创建并保存历史标签页引用
        self.history_tab = HistoryTab()
        tab_widget.addTab(self.history_tab, "上报历史")
        
        # 添加标签页
        tab_widget.addTab(main_tab, "主页面")
        tab_widget.addTab(ConfigTab(), "基本配置")
        tab_widget.addTab(APIConfigTab(), "接口配置")  # 添加API配置标签页
        tab_widget.addTab(TableMappingTab(), "字段映射")
        tab_widget.addTab(ExcelMappingTab(), "Excel映射")
        tab_widget.addTab(ImportTab(), "数据导入")
        
        layout.addWidget(tab_widget)
        
        # 初始化工作线程
        self.worker = None
        
    def log(self, message):
        """添加日志到显示区域"""
        self.log_display.append(message)
        
    def start_upload(self):
        """开始上报数据"""
        self.upload_button.setEnabled(False)
        self.log_display.clear()
        
        self.worker = WorkerThread()
        self.worker.update_signal.connect(self.log)
        self.worker.finished_signal.connect(self.handle_finished)
        self.worker.refresh_history_signal.connect(self.refresh_history)  # 连接刷新历史信号
        self.worker.start()
        
    def handle_finished(self, success, message):
        """处理上报完成"""
        self.log(message)
        self.upload_button.setEnabled(True)
        
        if success:
            QMessageBox.information(self, "成功", "数据上报成功！")
        else:
            QMessageBox.warning(self, "错误", "数据上报失败，请查看日志了解详情。")

    def refresh_history(self):
        """刷新历史记录"""
        if self.history_tab:
            self.history_tab.refresh_history()

class APIConfigTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.config_history_file = 'api_config_history.json'
        self.config_history = self.load_config_history()
        self.initUI()
        
    def initUI(self):
        layout = QVBoxLayout()
        
        # 添加配置选择下拉框
        config_select_layout = QHBoxLayout()
        self.config_combo = QComboBox()
        self.config_combo.currentIndexChanged.connect(self.on_config_selected)
        
        save_as_button = QPushButton("保存为新配置")
        save_as_button.clicked.connect(self.save_as_new_config)
        delete_button = QPushButton("删除当前配置")
        delete_button.clicked.connect(self.delete_current_config)
        
        config_select_layout.addWidget(QLabel("选择配置:"))
        config_select_layout.addWidget(self.config_combo)
        config_select_layout.addWidget(save_as_button)
        config_select_layout.addWidget(delete_button)
        
        layout.addLayout(config_select_layout)
        
        # 添加标题
        title = QLabel("API接口字段配置")
        title.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                margin: 10px 0;
            }
        """)
        layout.addWidget(title)
        
        # 导入配置区域
        import_group = QGroupBox("导入配置")
        import_layout = QVBoxLayout()
        
        # 文件选择区域
        file_select_layout = QHBoxLayout()
        self.file_path = QLineEdit()
        self.file_path.setReadOnly(True)
        self.file_path.setPlaceholderText("请选择Excel配置文件...")
        
        select_button = QPushButton("选择文件")
        select_button.clicked.connect(self.select_file)
        select_button.setStyleSheet('''
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        ''')
        
        file_select_layout.addWidget(self.file_path)
        file_select_layout.addWidget(select_button)
        import_layout.addLayout(file_select_layout)
        
        # 导入按钮
        import_button = QPushButton("导入配置")
        import_button.clicked.connect(self.import_config)
        import_button.setStyleSheet('''
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                margin-top: 10px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        ''')
        import_layout.addWidget(import_button)
        
        import_group.setLayout(import_layout)
        layout.addWidget(import_group)
        
        # 当前配置显示区域
        config_group = QGroupBox("当前配置")
        config_layout = QVBoxLayout()
        
        self.config_table = QTableWidget()
        self.config_table.setColumnCount(4)
        self.config_table.setHorizontalHeaderLabels(['字段名称', 'API字段名', '字段类型', '是否必填'])
        self.config_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        config_layout.addWidget(self.config_table)
        config_group.setLayout(config_layout)
        layout.addWidget(config_group)
        
        # 导出按钮
        export_button = QPushButton("导出当前配置")
        export_button.clicked.connect(self.export_config)
        export_button.setStyleSheet('''
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 8px;
                border: none;
                border-radius: 4px;
                margin-top: 10px;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
        ''')
        layout.addWidget(export_button)
        
        # 帮助说明
        help_text = QLabel("""
        说明：
        1. 通过Excel文件导入新的API接口字段配置
        2. Excel文件格式要求：
           - 必须包含：字段名称、API字段名、字段类型、是否必填 四列
           - 字段类型支持：string、int、float、date等
           - 是否必填填写：是/否
        3. 导入新配置后会自动更新系统中的字段映射关系
        """)
        help_text.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                padding: 10px;
                border-radius: 4px;
                border: 1px solid #dee2e6;
                color: #495057;
            }
        """)
        layout.addWidget(help_text)
        
        self.setLayout(layout)
        
        # 加载当前配置
        self.load_current_config()
        
        # 更新配置列表
        self.update_config_list()
        
    def select_file(self):
        """选择Excel文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择Excel配置文件",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.file_path.setText(file_path)
            
    def import_config(self):
        """导入配置"""
        if not self.file_path.text():
            QMessageBox.warning(self, "错误", "请先选择配置文件！")
            return
            
        try:
            # 读取Excel文件
            df = pd.read_excel(self.file_path.text())
            
            # 验证必要列
            required_columns = ['字段名称', 'API字段名', '字段类型', '是否必填']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                QMessageBox.warning(self, "错误", f"缺少必要列：{', '.join(missing_columns)}")
                return
                
            # 保存配置
            config = []
            for _, row in df.iterrows():
                config.append({
                    'name': row['字段名称'],
                    'api_field': row['API字段名'],
                    'type': row['字段类型'],
                    'required': row['是否必填'] == '是'
                })
                
            # 保存到配置文件
            with open('api_config.json', 'w', encoding='utf-8') as f:
                json.dump({'fields': config}, f, indent=4, ensure_ascii=False)
                
            # 刷新显示
            self.load_current_config()
            
            QMessageBox.information(self, "成功", "配置导入成功！")
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导入失败: {str(e)}")
            
    def load_current_config(self):
        """加载当前配置"""
        try:
            # 获取默认字段配置
            default_fields = self.get_default_fields()
            
            # 如果配置文件存在，读取配置
            if os.path.exists('api_config.json'):
                with open('api_config.json', 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    if not config.get('fields'):  # 如果没有字段配置
                        config['fields'] = default_fields
            else:
                # 如果配置文件不存在，使用默认配置
                config = {'fields': default_fields}
                # 保存默认配置到文件
                with open('api_config.json', 'w', encoding='utf-8') as f:
                    json.dump(config, f, indent=4, ensure_ascii=False)
            
            # 显示配置到表格
            self.display_config(config['fields'])
            
            # 更新配置列表
            if not any(c['name'] == '默认配置' for c in self.config_history['configurations']):
                self.config_history['configurations'].insert(0, {
                    'name': '默认配置',
                    'fields': default_fields
                })
                self.save_config_history()
                self.update_config_list()
                
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载配置失败: {str(e)}")
            print(f"加载配置失败: {str(e)}")  # 添加调试信息

    def export_config(self):
        """导出当前配置"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "导出配置",
                "api_config_export.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
                
            # 获取默认字段配置
            default_fields = self.get_default_fields()
            
            # 准备导出数据
            data = []
            for field in default_fields:
                data.append({
                    '字段名称': field['name'],
                    'API字段名': field['api_field'],
                    '字段类型': field['type'],
                    '是否必填': '是' if field['required'] else '否',
                    '说明': field.get('description', '')
                })
                
            # 导出到Excel
            df = pd.DataFrame(data)
            
            # 使用ExcelWriter以便设置格式
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='API字段配置')
                
                # 获取工作表
                worksheet = writer.sheets['API字段配置']
                
                # 调整列宽
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(col)
                    )
                    worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2
                
                # 设置表头样式
                header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                header_font = Font(bold=True)
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    
                # 设置必填字段的样式
                red_font = Font(color='FF0000')
                for row in range(2, worksheet.max_row + 1):
                    if worksheet.cell(row=row, column=4).value == '是':
                        worksheet.cell(row=row, column=4).font = red_font
                        
            QMessageBox.information(self, "成功", f"配置已导出到：\n{file_path}")
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出失败: {str(e)}")

    def load_config_history(self):
        """加载配置历史"""
        try:
            if os.path.exists(self.config_history_file):
                with open(self.config_history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return {'configurations': []}
        except Exception:
            return {'configurations': []}
            
    def save_config_history(self):
        """保存配置历史"""
        try:
            with open(self.config_history_file, 'w', encoding='utf-8') as f:
                json.dump(self.config_history, f, indent=4, ensure_ascii=False)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存配置历史失败: {str(e)}")
            
    def update_config_list(self):
        """更新配置列表"""
        self.config_combo.clear()
        
        # 确保默认配置存在
        if not any(config['name'] == '默认配置' for config in self.config_history['configurations']):
            self.config_history['configurations'].insert(0, {
                'name': '默认配置',
                'fields': self.get_default_fields()
            })
            self.save_config_history()
            
        # 更新下拉框
        for config in self.config_history['configurations']:
            self.config_combo.addItem(config['name'])
            
    def on_config_selected(self, index):
        """配置选择改变时的处理"""
        if index < 0:
            return
            
        config_name = self.config_combo.currentText()
        
        # 如果是默认配置，禁用删除按钮
        delete_button = self.findChild(QPushButton, "delete_config_button")
        if delete_button:
            delete_button.setEnabled(config_name != '默认配置')
            
        # 加载选中的配置
        for config in self.config_history['configurations']:
            if config['name'] == config_name:
                self.display_config(config['fields'])
                break
                
    def save_as_new_config(self):
        """保存为新配置"""
        name, ok = QInputDialog.getText(self, '保存配置', '请输入配置名称:')
        if ok and name:
            if name == '默认配置':
                QMessageBox.warning(self, "错误", "不能使用'默认配置'作为名称！")
                return
                
            if name in [self.config_combo.itemText(i) for i in range(self.config_combo.count())]:
                QMessageBox.warning(self, "错误", "配置名称已存在！")
                return
                
            # 收集当前配置
            fields = []
            for row in range(self.config_table.rowCount()):
                field = {
                    'name': self.config_table.item(row, 0).text(),
                    'api_field': self.config_table.item(row, 1).text(),
                    'type': self.config_table.item(row, 2).text(),
                    'required': self.config_table.item(row, 3).text() == '是'
                }
                fields.append(field)
                
            # 保存配置
            self.config_history['configurations'].append({
                'name': name,
                'fields': fields
            })
            self.save_config_history()
            
            # 更新列表并选择新配置
            self.update_config_list()
            self.config_combo.setCurrentText(name)
            
            QMessageBox.information(self, "成功", "配置保存成功！")
            
    def delete_current_config(self):
        """删除当前配置"""
        current_config = self.config_combo.currentText()
        if current_config == '默认配置':
            QMessageBox.warning(self, "错误", "不能删除默认配置！")
            return
            
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定要删除配置 '{current_config}' 吗？此操作不可恢复。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.config_history['configurations'] = [
                config for config in self.config_history['configurations']
                if config['name'] != current_config
            ]
            self.save_config_history()
            self.update_config_list()
            
    def get_default_fields(self):
        """获取默认字段配置"""
        return [
            {
                'name': '统一社会信用代码',
                'api_field': 'socialCreditCode',
                'type': 'string',
                'required': True
            },
            {
                'name': '企业名称',
                'api_field': 'compName',
                'type': 'string',
                'required': True
            },
            {
                'name': '零售点编码',
                'api_field': 'retailStoreCode',
                'type': 'string',
                'required': True
            },
            {
                'name': '零售点名称',
                'api_field': 'retailStoreName',
                'type': 'string',
                'required': True
            },
            {
                'name': '上报日期',
                'api_field': 'reportDate',
                'type': 'date',
                'required': True
            },
            {
                'name': '商品编码',
                'api_field': 'selfCommondityCode',
                'type': 'string',
                'required': True
            },
            {
                'name': '商品名称',
                'api_field': 'selfCommondityName',
                'type': 'string',
                'required': True
            },
            {
                'name': '单位',
                'api_field': 'unit',
                'type': 'string',
                'required': True
            },
            {
                'name': '规格',
                'api_field': 'spec',
                'type': 'string',
                'required': True
            },
            {
                'name': '条码',
                'api_field': 'barcode',
                'type': 'string',
                'required': True
            },
            {
                'name': '数据类型',
                'api_field': 'dataType',
                'type': 'int',
                'required': True,
                'description': '1期初库存、2入库量、3销售量、4价格'
            },
            {
                'name': '数据值',
                'api_field': 'dataValue',
                'type': 'float',
                'required': True
            },
            {
                'name': '转换标志',
                'api_field': 'dataConvertFlag',
                'type': 'int',
                'required': True,
                'description': '默认值2'
            },
            {
                'name': '供应商编码',
                'api_field': 'supplierCode',
                'type': 'string',
                'required': True
            },
            {
                'name': '供应商名称',
                'api_field': 'supplierName',
                'type': 'string',
                'required': True
            },
            {
                'name': '生产商名称',
                'api_field': 'manufatureName',
                'type': 'string',
                'required': True
            },
            {
                'name': '产地编码',
                'api_field': 'originCode',
                'type': 'string',
                'required': True
            },
            {
                'name': '产地名称',
                'api_field': 'originName',
                'type': 'string',
                'required': True
            },
            {
                'name': '场景标志',
                'api_field': 'sceneflag',
                'type': 'int',
                'required': True,
                'description': '默认值1'
            }
        ]

    def display_config(self, fields):
        """显示配置到表格"""
        try:
            # 清空表格
            self.config_table.setRowCount(0)
            
            # 获取默认字段配置
            default_fields = self.get_default_fields()
            
            # 如果没有传入字段，使用默认字段
            if not fields:
                fields = default_fields
                
            # 确保所有必要字段都存在
            field_dict = {field['api_field']: field for field in fields}
            for default_field in default_fields:
                if default_field['api_field'] not in field_dict:
                    field_dict[default_field['api_field']] = default_field
                    
            # 显示配置
            self.config_table.setRowCount(len(field_dict))
            for i, field in enumerate(field_dict.values()):
                # 字段名称
                name_item = QTableWidgetItem(field['name'])
                self.config_table.setItem(i, 0, name_item)
                
                # API字段名
                api_field_item = QTableWidgetItem(field['api_field'])
                self.config_table.setItem(i, 1, api_field_item)
                
                # 字段类型
                type_item = QTableWidgetItem(field.get('type', 'string'))
                self.config_table.setItem(i, 2, type_item)
                
                # 是否必填
                required = field.get('required', True)
                required_item = QTableWidgetItem('是' if required else '否')
                required_item.setForeground(Qt.red if required else Qt.black)
                self.config_table.setItem(i, 3, required_item)
                
                # 如果有字段说明，添加到工具提示
                if 'description' in field:
                    for col in range(4):
                        item = self.config_table.item(i, col)
                        if item:
                            item.setToolTip(field['description'])
                        
            # 调整列宽
            self.config_table.resizeColumnsToContents()
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"显示配置失败: {str(e)}")
            print(f"显示配置失败: {str(e)}")  # 添加调试信息

def main():
    app = QApplication(sys.argv)
    
    # 设置应用样式
    app.setStyle('Fusion')
    
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main() 